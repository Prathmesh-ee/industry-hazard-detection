"""
Industry Hazard Detection — Real-Time Dashboard
------------------------------------------------
Requirements:  pip install pyserial matplotlib openpyxl
Usage:         python dashboard.py
"""

import serial
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.animation as animation
from collections import deque
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# ─── CONFIG ───────────────────────────────────────────────────────────────────
COM_PORT   = "COM7"
BAUD_RATE  = 9600
MAX_POINTS = 60
LOG_FILE   = "hazard_log.xlsx"
# ──────────────────────────────────────────────────────────────────────────────

# Data buffers for graphs
temps      = deque([None] * MAX_POINTS, maxlen=MAX_POINTS)
hums       = deque([None] * MAX_POINTS, maxlen=MAX_POINTS)
timestamps = deque([None] * MAX_POINTS, maxlen=MAX_POINTS)

# Latest sensor states
latest = {"FIRE": 0, "SMOKE": 0, "LDR": 0, "TEMP": 0, "HUM": 0}

# Previous state for change detection
prev_state = {"FIRE": -1, "SMOKE": -1, "LDR": -1,
              "TEMP": -1, "HUM": -1}

# ─── EXCEL SETUP ──────────────────────────────────────────────────────────────
session_name = datetime.now().strftime("Session %d-%m %H-%M")
if os.path.exists(LOG_FILE):
    wb = load_workbook(LOG_FILE)
else:
    wb = Workbook()
    wb.remove(wb.active)

ws = wb.create_sheet(session_name)
ws.append(["Timestamp", "Temp(C)", "Humidity(%)", "Fire", "Smoke", "LDR", "Reason"])
wb.save(LOG_FILE)

# ─── SERIAL SETUP ─────────────────────────────────────────────────────────────
try:
    ser = serial.Serial(COM_PORT, BAUD_RATE, timeout=1)
    print(f"[OK] Connected to {COM_PORT}")
except Exception as e:
    print(f"[ERROR] Could not open serial port: {e}")
    exit()

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def parse_line(line):
    try:
        parts = line.strip().split(",")
        data = {}
        for p in parts:
            k, v = p.split(":")
            data[k.strip()] = float(v.strip())
        return data
    except:
        return None

def get_alert_level():
    if latest["FIRE"] or latest["SMOKE"]:
        return "CRITICAL", "#ff2222"
    if latest["TEMP"] > 45:
        return "WARNING", "#ffaa00"
    return "SAFE", "#00cc66"

# ─── FIGURE SETUP ─────────────────────────────────────────────────────────────
plt.style.use("dark_background")
fig = plt.figure(figsize=(14, 8), facecolor="#0d0d0d")
fig.suptitle("INDUSTRY HAZARD MONITORING SYSTEM",
             fontsize=15, color="#ffffff", fontweight="bold", y=0.98)

ax_temp   = fig.add_subplot(2, 2, 1)
ax_hum    = fig.add_subplot(2, 2, 2)
ax_status = fig.add_subplot(2, 1, 2)

plt.subplots_adjust(hspace=0.45, wspace=0.35,
                    left=0.07, right=0.97, top=0.92, bottom=0.08)

def style_ax(ax, title, ylabel, color):
    ax.set_facecolor("#1a1a1a")
    ax.set_title(title, color=color, fontsize=11, pad=6)
    ax.set_ylabel(ylabel, color="#aaaaaa", fontsize=9)
    ax.set_xlabel("Time", color="#aaaaaa", fontsize=9)
    ax.tick_params(colors="#666666", labelsize=7)
    for spine in ax.spines.values():
        spine.set_edgecolor("#333333")

# ─── ANIMATE ──────────────────────────────────────────────────────────────────
def animate(frame):
    global prev_state

    # Read serial
    try:
        raw = ser.readline().decode("utf-8", errors="ignore")
    except:
        return

    data = parse_line(raw)
    if not data:
        return

    # Skip bad DHT readings
    t_val = data.get("TEMP", 0)
    h_val = data.get("HUM", 0)
    if t_val == 0 or h_val == 0:
        return

    now = datetime.now().strftime("%H:%M:%S")

    # Update latest values
    latest["TEMP"]  = t_val
    latest["HUM"]   = h_val
    latest["FIRE"]  = int(data.get("FIRE",  0))
    latest["SMOKE"] = int(data.get("SMOKE", 0))
    latest["LDR"]   = int(data.get("LDR",   0))

    # Update graph buffers
    temps.append(t_val)
    hums.append(h_val)
    timestamps.append(now)

    # ── Change detection and logging ──
    temp_rounded = round(t_val, 1)
    hum_rounded  = round(h_val, 1)

    changed = (
        latest["FIRE"]  != prev_state["FIRE"]  or
        latest["SMOKE"] != prev_state["SMOKE"] or
        latest["LDR"]   != prev_state["LDR"]   or
        temp_rounded    != prev_state["TEMP"]  or
        hum_rounded     != prev_state["HUM"]
    )

    if changed:
        reasons = []
        if latest["FIRE"]  != prev_state["FIRE"]:
            reasons.append("FIRE DETECTED"   if latest["FIRE"]  else "FIRE CLEARED")
        if latest["SMOKE"] != prev_state["SMOKE"]:
            reasons.append("SMOKE DETECTED"  if latest["SMOKE"] else "SMOKE CLEARED")
        if latest["LDR"]   != prev_state["LDR"]:
            reasons.append("NIGHT"           if latest["LDR"]   else "DAY")
        if temp_rounded    != prev_state["TEMP"]:
            reasons.append(f"TEMP {temp_rounded}C")
        if hum_rounded     != prev_state["HUM"]:
            reasons.append(f"HUM {hum_rounded}%")

        ws.append([now, t_val, h_val,
                   latest["FIRE"], latest["SMOKE"], latest["LDR"],
                   " | ".join(reasons)])
        wb.save(LOG_FILE)

        # Update prev state
        prev_state["FIRE"]  = latest["FIRE"]
        prev_state["SMOKE"] = latest["SMOKE"]
        prev_state["LDR"]   = latest["LDR"]
        prev_state["TEMP"]  = temp_rounded
        prev_state["HUM"]   = hum_rounded

    # ── Build graph data ──
    valid = [(t, v1, v2) for t, v1, v2
             in zip(timestamps, temps, hums) if t is not None]
    if not valid:
        return

    t_labels = [v[0] for v in valid]
    t_vals   = [v[1] for v in valid]
    h_vals   = [v[2] for v in valid]
    x        = range(len(valid))

    tick_count     = min(5, len(t_labels))
    tick_positions = [int(i * (len(t_labels) - 1) / (tick_count - 1))
                      for i in range(tick_count)] if tick_count > 1 else [0]
    tick_labels    = [t_labels[i] for i in tick_positions]

    # ── Temperature plot ──
    ax_temp.cla()
    style_ax(ax_temp, "Temperature (°C)", "°C", "#ff6b6b")
    ax_temp.plot(x, t_vals, color="#ff6b6b", linewidth=2)
    ax_temp.fill_between(x, t_vals, alpha=0.15, color="#ff6b6b")
    ax_temp.axhline(40, color="#ff4444", linestyle="--", linewidth=1, alpha=0.6)
    ax_temp.text(1, 41, "threshold", color="#ff4444", fontsize=7)
    ax_temp.set_ylim(0, 80)
    ax_temp.set_xticks(tick_positions)
    ax_temp.set_xticklabels(tick_labels, rotation=30, ha="right")

    # ── Humidity plot ──
    ax_hum.cla()
    style_ax(ax_hum, "Humidity (%)", "%", "#4ecdc4")
    ax_hum.plot(x, h_vals, color="#4ecdc4", linewidth=2)
    ax_hum.fill_between(x, h_vals, alpha=0.15, color="#4ecdc4")
    ax_hum.set_ylim(0, 100)
    ax_hum.set_xticks(tick_positions)
    ax_hum.set_xticklabels(tick_labels, rotation=30, ha="right")

    # ── Status panel ──
    ax_status.cla()
    ax_status.set_facecolor("#1a1a1a")
    ax_status.axis("off")

    alert_label, alert_color = get_alert_level()

    # Alert banner
    ax_status.add_patch(mpatches.Rectangle(
        (0, 0.55), 1, 0.45,
        transform=ax_status.transAxes,
        color=alert_color, alpha=0.15))
    ax_status.text(0.5, 0.75, f"  {alert_label}",
                   transform=ax_status.transAxes,
                   ha="center", va="center",
                   fontsize=22, fontweight="bold", color=alert_color)

    # Sensor tiles
    sensors = [
        ("FIRE",               latest["FIRE"],  "#ff4444"),
        ("SMOKE",              latest["SMOKE"], "#ffaa00"),
        ("NIGHT",              latest["LDR"],   "#8888ff"),
        (f"TEMP: {t_val:.1f} C", 0,            "#ff6b6b"),
        (f"HUM:  {h_val:.1f} %", 0,            "#4ecdc4"),
    ]

    for i, (label, active, color) in enumerate(sensors):
        x_pos      = 0.05 + i * 0.19
        bg_color   = color     if active else "#2a2a2a"
        text_color = "#ffffff" if active else "#888888"

        ax_status.add_patch(mpatches.FancyBboxPatch(
            (x_pos, 0.05), 0.16, 0.38,
            transform=ax_status.transAxes,
            boxstyle="round,pad=0.02",
            facecolor=bg_color, alpha=0.3 if active else 0.15,
            edgecolor=color, linewidth=1.5))

        ax_status.text(x_pos + 0.08, 0.24, label,
                       transform=ax_status.transAxes,
                       ha="center", va="center",
                       fontsize=10, color=text_color, fontweight="bold")

    ax_status.text(0.98, 0.01,
                   f"Last update: {now}  |  Logging -> {LOG_FILE}",
                   transform=ax_status.transAxes,
                   ha="right", va="bottom", fontsize=8, color="#555555")

ani = animation.FuncAnimation(fig, animate, interval=1000, cache_frame_data=False)

print("[INFO] Dashboard running. Close the window to stop.")
print(f"[INFO] Logging anomalies -> {LOG_FILE}")

plt.show()
wb.save(LOG_FILE)
